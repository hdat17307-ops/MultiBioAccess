"""Chuyen doi data/employees.json <-> data/employees.xlsx de tien chinh sua bang Excel."""

import argparse
import json
from pathlib import Path

from openpyxl import Workbook, load_workbook

from config import EMPLOYEE_DB_PATH, EMPLOYEE_XLSX_PATH

JSON_PATH = Path(EMPLOYEE_DB_PATH)
XLSX_PATH = Path(EMPLOYEE_XLSX_PATH)
COLUMNS = ["id", "name", "department", "face_image", "card_id"]


def export_to_excel():
    with open(JSON_PATH, "r", encoding="utf-8") as f:
        data = json.load(f)

    wb = Workbook()
    ws = wb.active
    ws.title = "employees"
    ws.append(COLUMNS)
    for emp in data["employees"]:
        ws.append([emp.get(col, "") for col in COLUMNS])

    wb.save(XLSX_PATH)
    print(f"Da xuat {len(data['employees'])} nhan vien sang {XLSX_PATH}")


def import_from_excel():
    wb = load_workbook(XLSX_PATH)
    ws = wb["employees"]

    employees = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        employees.append({col: row[i] for i, col in enumerate(COLUMNS)})

    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump({"employees": employees}, f, indent=2, ensure_ascii=False)
        f.write("\n")

    print(f"Da nhap {len(employees)} nhan vien tu {XLSX_PATH} vao {JSON_PATH}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    subparsers = parser.add_subparsers(dest="command", required=True)
    subparsers.add_parser("export", help="Xuat employees.json sang employees.xlsx")
    subparsers.add_parser("import", help="Nhap employees.xlsx vao employees.json")
    args = parser.parse_args()

    if args.command == "export":
        export_to_excel()
    else:
        import_from_excel()
